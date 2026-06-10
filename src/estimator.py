import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import date
import price_master


def estimate_from_quantities(quantities: dict) -> list[dict]:
    """
    チェック結果の数量から積算明細を生成する。
    数量が取れていない項目は坪単価方式でフォールバック。
    """
    items = []
    floor_area = quantities.get("floor_area")
    window_count = quantities.get("window_count") or 0
    door_count = quantities.get("door_count") or 0
    outer_wall_length = quantities.get("outer_wall_length")

    # 外壁面積の概算（外壁周長 × 平均2.8m高さ）
    outer_wall_area = (outer_wall_length * 2.8) if outer_wall_length else None

    # 基礎工事
    if floor_area:
        items.append({
            "工種": "基礎工事",
            "品目": "ベタ基礎",
            "数量": floor_area,
            "単位": "m2",
            "単価": price_master.lookup("基礎工事", "ベタ基礎") or 25000,
        })

    # 木工事（床組）
    if floor_area:
        items.append({
            "工種": "木工事",
            "品目": "床組",
            "数量": floor_area,
            "単位": "m2",
            "単価": price_master.lookup("木工事", "床組") or 8000,
        })

    # 外壁工事
    if outer_wall_area:
        items.append({
            "工種": "外壁工事",
            "品目": "サイディング",
            "数量": round(outer_wall_area, 1),
            "単位": "m2",
            "単価": price_master.lookup("外壁工事", "サイディング") or 7000,
        })
        items.append({
            "工種": "仮設工事",
            "品目": "足場設置",
            "数量": round(outer_wall_area, 1),
            "単位": "m2",
            "単価": price_master.lookup("仮設工事", "足場設置") or 1500,
        })

    # 屋根工事
    if floor_area:
        roof_area = round(floor_area * 1.3, 1)  # 勾配係数1.3で概算
        items.append({
            "工種": "屋根工事",
            "品目": "瓦葺き",
            "数量": roof_area,
            "単位": "m2",
            "単価": price_master.lookup("屋根工事", "瓦葺き") or 12000,
        })

    # 建具工事
    if window_count > 0:
        items.append({
            "工種": "建具工事",
            "品目": "引き違い窓（大）",
            "数量": window_count,
            "単位": "箇所",
            "単価": price_master.lookup("建具工事", "引き違い窓（大）") or 50000,
        })
    if door_count > 0:
        items.append({
            "工種": "建具工事",
            "品目": "室内ドア",
            "数量": door_count,
            "単位": "箇所",
            "単価": price_master.lookup("建具工事", "室内ドア") or 40000,
        })

    # 内装工事
    if floor_area:
        items.append({
            "工種": "内装工事",
            "品目": "床フローリング",
            "数量": floor_area,
            "単位": "m2",
            "単価": price_master.lookup("内装工事", "床フローリング") or 8000,
        })
        wall_area = round(floor_area * 3.5, 1)  # 壁面積の概算
        items.append({
            "工種": "内装工事",
            "品目": "クロス仕上げ（壁）",
            "数量": wall_area,
            "単位": "m2",
            "単価": price_master.lookup("内装工事", "クロス仕上げ（壁）") or 1200,
        })

    # 設備工事（固定費）
    items.append({
        "工種": "設備工事",
        "品目": "給排水配管",
        "数量": 1,
        "単位": "式",
        "単価": price_master.lookup("設備工事", "給排水配管") or 500000,
    })
    items.append({
        "工種": "設備工事",
        "品目": "電気配線",
        "数量": 1,
        "単位": "式",
        "単価": price_master.lookup("設備工事", "電気配線") or 400000,
    })

    # 金額計算
    for item in items:
        item["金額"] = item["数量"] * item["単価"]

    return items


def export_excel(items: list[dict], output_path: Path, project_name: str = "物件名") -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "見積書"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    category_fill = PatternFill(fill_type="solid", fgColor="D6E4F7")
    category_font = Font(bold=True, size=10)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # タイトル
    ws.merge_cells("A1:G1")
    ws["A1"] = f"概算見積書　― {project_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"作成日: {date.today().isoformat()}"
    ws["A2"].font = Font(size=9, color="666666")
    ws["G2"] = "※ 本見積はAIによる概算です"
    ws["G2"].font = Font(size=9, color="AA0000", italic=True)
    ws["G2"].alignment = Alignment(horizontal="right")

    # ヘッダー行
    headers = ["No.", "工種", "品目", "数量", "単位", "単価（円）", "金額（円）"]
    widths = [5, 15, 20, 10, 8, 14, 16]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    # データ行
    df = pd.DataFrame(items)
    row = 5
    prev_category = None
    subtotal_rows: list[tuple[str, int, int]] = []  # (category, start_row, end_row)
    cat_start = row

    for i, r in df.iterrows():
        if r["工種"] != prev_category:
            if prev_category is not None:
                subtotal_rows.append((prev_category, cat_start, row - 1))
            prev_category = r["工種"]
            cat_start = row

        cells_data = [i + 1, r["工種"], r["品目"], r["数量"], r["単位"],
                      r["単価"], r["金額"]]
        for col, val in enumerate(cells_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.font = Font(size=10)
            if col in (4, 6, 7):
                cell.alignment = Alignment(horizontal="right")
                if col in (6, 7):
                    cell.number_format = "#,##0"
        row += 1

    if prev_category is not None:
        subtotal_rows.append((prev_category, cat_start, row - 1))

    # 合計行
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "合　計"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    ws[f"A{row}"].alignment = Alignment(horizontal="right")
    total = sum(item["金額"] for item in items)
    ws[f"G{row}"] = total
    ws[f"G{row}"].font = Font(bold=True, size=11)
    ws[f"G{row}"].number_format = "#,##0"
    ws[f"G{row}"].alignment = Alignment(horizontal="right")
    for col in range(1, 8):
        ws.cell(row=row, column=col).fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        ws.cell(row=row, column=col).border = border

    # 消費税行
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "消費税（10%）"
    ws[f"A{row}"].alignment = Alignment(horizontal="right")
    ws[f"A{row}"].font = Font(size=10)
    tax = int(total * 0.1)
    ws[f"G{row}"] = tax
    ws[f"G{row}"].number_format = "#,##0"
    ws[f"G{row}"].alignment = Alignment(horizontal="right")

    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "税込合計"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="right")
    ws[f"G{row}"] = total + tax
    ws[f"G{row}"].font = Font(bold=True, size=12)
    ws[f"G{row}"].number_format = "#,##0"
    ws[f"G{row}"].alignment = Alignment(horizontal="right")
    for col in range(1, 8):
        ws.cell(row=row, column=col).fill = PatternFill(fill_type="solid", fgColor="FFD700")
        ws.cell(row=row, column=col).border = border

    ws.freeze_panes = "A5"
    wb.save(output_path)
    return output_path
